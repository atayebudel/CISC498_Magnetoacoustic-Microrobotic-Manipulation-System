"""
GUI Functions Module for Magnetoacoustic Microrobotic Manipulation System.

This module implements the main application window and handles all GUI interactions
for controlling and tracking microrobots using magnetic and acoustic fields.

Classes:
    MainWindow: Main application window with video tracking, control, and data recording.
"""

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QFileDialog
import sys
from PyQt5.QtGui import QWheelEvent
from PyQt5 import QtGui
from PyQt5.QtWidgets import QWidget
from PyQt5.QtGui import QPixmap,QIcon
from PyQt5.QtCore import Qt, QTimer, PYQT_VERSION_STR, QEvent
from PyQt5 import QtWidgets, QtGui, QtCore
import queue
import cv2
import os
from os.path import expanduser
import openpyxl 
import pandas as pd
from datetime import datetime
import sys
from PyQt5.QtWidgets import QApplication
import numpy as np
import cv2
import matplotlib.pyplot as plt 
import time
import platform
from serial.tools import list_ports
os.environ["SDL_JOYSTICK_ALLOW_BACKGROUND_EVENTS"] = "1"
os.environ['PYGAME_HIDE_SUPPORT_PROMPT'] = "hide"
import pygame
try:
    import EasyPySpin
except Exception:
    pass

from classes.gui_widgets import Ui_MainWindow
from classes.arduino_class import ArduinoHandler
from classes.joystick_class import Mac_Joystick,Linux_Joystick,Windows_Joystick
from classes.simulation_class import HelmholtzSimulator
from classes.control_class import Controller
from classes.path_planning_class import Path_Planner



from classes import tracking_panel

class MainWindow(QtWidgets.QMainWindow):
    """
    Main application window for the Magnetoacoustic Microrobotic Manipulation System.
    
    This class handles:
    - Video capture and display from FLIR cameras or video files
    - Real-time microrobot tracking and visualization
    - Magnetic field control via Arduino
    - Joystick input for manual control
    - Path planning and autonomous navigation
    - Data recording and export to Excel
    - Simulation of magnetic field configurations
    
    Signals:
        positionChanged (QPoint): Emitted when mouse position changes on video feed.
    
    Attributes:
        ui (Ui_MainWindow): UI widgets and layout
        arduino (ArduinoHandler): Interface to Arduino hardware controller
        tracker (tracking_panel): Video tracking and analysis
        simulator (HelmholtzSimulator): Magnetic field visualization
        control_robot (Controller): Autonomous control algorithms
        path_planner (Path_Planner): Path planning algorithms
    """
    positionChanged = QtCore.pyqtSignal(QtCore.QPoint)

    def __init__(self, parent=None):
        """
        Initialize the main window and all subsystems.
        
        Args:
            parent (QWidget, optional): Parent widget. Defaults to None.
        """
        super(MainWindow, self).__init__(parent=parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        

        
        
        #self.showMaximized()

        #resize some widgets to fit the screen better
        screen  = QtWidgets.QDesktopWidget().screenGeometry(-1)
        
        self.window_width = screen.width()
        self.window_height = screen.height()
        self.resize(self.window_width, self.window_height)
        self.display_width = self.window_width# self.ui.frameGeometry().width()

        self.displayheightratio = 0.79
        self.framesliderheightratio = 0.031
        self.textheightratio = .129
        self.tabheightratio = 0.925
        self.tabheightratio = 0.925
        
        self.aspectratio = 1041/801
        self.resize_widgets()

    
      
        #create folder in homerdiractory of user
        if "Windows" in platform.platform():
            home_dir = expanduser("D:")
            new_dir_name = "Tracking Data"
            desktop_path = os.path.join(home_dir, "Microrobots")
            self.new_dir_path = os.path.join(desktop_path, new_dir_name)
            if not os.path.exists(self.new_dir_path):
                os.makedirs(self.new_dir_path)
        else:
            #home_dir = expanduser("~")
            #new_dir_name = "Tracking Data"
            #desktop_path = os.path.join(home_dir, "Desktop")
            
            self.new_dir_path = "Data"#os.path.join(desktop_path, new_dir_name)
            if not os.path.exists(self.new_dir_path):
                os.makedirs(self.new_dir_path)



        self.zoom_x, self.zoom_y, self.zoomscale, self.scrollamount = 1,0,0,0
        self.croppedresult = None
        self.frame_number = 0
        self.robots = []
        self.cells = []
        self.videopath = 0
        self.cap = None
        self.tracker = None
        self.populate_serial_ports()
        self.arduino = None

        #record variables
        self.recorder = None
        self.frame_queue = queue.Queue(maxsize=100)  # make a queue to store frames in for the recording feature
        self.output_file_name = datetime.now().strftime("%Y-%m-%d_%H-%M-%S_%f")
   

        self.save_status = False
        self.output_workbook = None
        self.ricochet_counter_x = [0]
        self.ricochet_counter_y = [0]
        
        
        self.drawing = False
        self.acoustic_frequency = 0
        self.autoacousticstatus = False
        self.gradient_status = 0
        self.equal_field_status = 0
        self.magnetic_field_list = []
        
        self.actions = [0,0,0,0,0,0,0,0,0,0,0,0,0]
        self.Bx, self.By, self.Bz = 0,0,0
        self.Mx, self.My, self.Mz = 0,0,0
        self.alpha, self.gamma, self.psi, self.freq = 0,0,0,0
        self.field_magnitude = 100

      
        #control tab functions
        self.path_planner_status = False
        self.control_status = False
        self.joystick_status = False
        self.manual_status = False


  
        if "mac" in platform.platform():
            self.tbprint("Detected OS: macos")
            self.joystick_actions = Mac_Joystick()
        elif "Linux" in platform.platform():
            self.tbprint("Detected OS: Linux")
            self.joystick_actions = Linux_Joystick()
        elif "Windows" in platform.platform():
            self.tbprint("Detected OS:  Windows")
            self.joystick_actions = Windows_Joystick()
        else:
            self.tbprint("undetected operating system")
        

        
        
        #define, simulator class, pojection class, and acoustic class
        self.simulator = HelmholtzSimulator(self.ui.magneticfieldsimlabel, width=310, height=310, dpi=200)

        
        #make instance of algorithm class both control and path planning
        self.control_robot = Controller()
        self.path_planner = Path_Planner()
        



        self.setFile()
        
        pygame.init()
        if pygame.joystick.get_count() == 0:
            self.tbprint("No Joystick Connected...")
            
        else:
            self.joystick = pygame.joystick.Joystick(0)
            self.joystick.init()
            self.tbprint("Connected to: "+str(self.joystick.get_name()))
        
      
        self.sensorupdatetimer = QTimer(self)
        self.sensorupdatetimer.timeout.connect(self.update_sensor_label)
        self.sensorupdatetimer.start(25)  # Update every 500 ms
        self.bx_sensor = 0
        self.by_sensor = 0 
        self.bz_sensor = 0

     

        #tracker tab functions
        self.ui.pausebutton.hide()
        self.ui.leftbutton.hide()
        self.ui.rightbutton.hide()
        self.ui.choosevideobutton.clicked.connect(self.selectFile)
        self.ui.startbutton.clicked.connect(self.start)
        self.ui.pausebutton.clicked.connect(self.pause)
        self.ui.rightbutton.clicked.connect(self.frameright)
        self.ui.leftbutton.clicked.connect(self.frameleft)
        self.ui.maskbutton.clicked.connect(self.showmask)
        self.ui.maskinvert_checkBox.toggled.connect(self.invertmaskcommand)
        self.ui.robotmasklowerbox.valueChanged.connect(self.get_slider_vals)
        self.ui.robotmaskupperbox.valueChanged.connect(self.get_slider_vals)
        self.ui.robotmaskdilationbox.valueChanged.connect(self.get_slider_vals)
        self.ui.robotmaskblurbox.valueChanged.connect(self.get_slider_vals)
        self.ui.robotcroplengthbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellmasklowerbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellmaskupperbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellmaskdilationbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellmaskblurbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellcroplengthbox.valueChanged.connect(self.get_slider_vals)
        self.ui.gradient_status_checkbox.toggled.connect(self.gradientcommand)
        self.ui.equal_field_checkbox.toggled.connect(self.equalfieldcommand)
        self.ui.savedatabutton.clicked.connect(self.savedata)
        self.ui.VideoFeedLabel.installEventFilter(self)
        self.ui.recordbutton.clicked.connect(self.toggle_recording)
        self.ui.controlbutton.clicked.connect(self.toggle_control_status)
        self.ui.memorybox.valueChanged.connect(self.get_slider_vals)
        self.ui.RRTtreesizebox.valueChanged.connect(self.get_slider_vals)
        self.ui.arrivalthreshbox.valueChanged.connect(self.get_slider_vals)
        self.ui.magneticfrequencydial.valueChanged.connect(self.get_slider_vals)
        self.ui.gammadial.valueChanged.connect(self.get_slider_vals)
        self.ui.psidial.valueChanged.connect(self.get_slider_vals)
        self.ui.applyacousticbutton.clicked.connect(self.apply_acoustic)
        self.ui.acousticfreq_spinBox.valueChanged.connect(self.get_acoustic_frequency)
        self.ui.alphaspinBox.valueChanged.connect(self.spinbox_alphachanged)
        self.ui.alphadial.valueChanged.connect(self.dial_alphachanged)
        self.ui.resetdefaultbutton.clicked.connect(self.resetparams)
        self.ui.simulationbutton.clicked.connect(self.toggle_simulation)

        self.ui.objectivebox.valueChanged.connect(self.get_objective)
        self.ui.exposurebox.valueChanged.connect(self.get_exposure)
        self.ui.joystickbutton.clicked.connect(self.toggle_joystick_status)
        self.ui.autoacousticbutton.clicked.connect(self.toggle_autoacoustic)
        self.ui.manualapplybutton.clicked.connect(self.get_manual_bfieldbuttons)
        self.ui.manualfieldBx.valueChanged.connect(self.get_slider_vals)
        self.ui.manualfieldBy.valueChanged.connect(self.get_slider_vals)
        self.ui.manualfieldBz.valueChanged.connect(self.get_slider_vals)
        self.ui.croppedmasktoggle.clicked.connect(self.showcroppedoriginal)
        self.ui.croppedrecordbutton.clicked.connect(self.croppedrecordfunction)
        self.ui.import_excel_actions.clicked.connect(self.read_excel_actions)
        self.ui.apply_actions.clicked.connect(self.apply_excel_actions)
        self.ui.arduino_portbox.currentTextChanged.connect(self.handle_port_change)

        self.ui.cleartrackingbutton.clicked.connect(self.clear_tracking)

        self.excel_file_name = None
        self.excel_actions_df = None
        self.excel_actions_status = False


        self.ui.make_inf_path.clicked.connect(self.makeinf_trajectory)

        

    def clear_tracking(self):
        """
        Clear all tracked robots and cells from the current session.
        
        Removes all tracking data including trajectories and magnetic field history.
        """
        if self.tracker is not None:
            del self.tracker.robot_list[:]
            del self.tracker.cell_list[:]
            del self.magnetic_field_list[:]
            del self.robots[:]
            del self.cells[:]
            self.apply_actions(False)
    



        
    def makeinf_trajectory(self):
        """
        Generate an infinity symbol (lemniscate) trajectory for the last tracked robot.
        
        Creates a smooth figure-eight path centered on the video frame using
        parametric equations. The trajectory size is controlled by the infinity_size
        spin box in the UI.
        """
        if self.tracker is not None:
            if len(self.tracker.robot_list)>0:
             


                a = self.ui.infinity_size.value()  # Controls the size
                center_x = self.video_width // 2
                center_y = self.video_height // 2

                # Generate points using parametric equations for a lemniscate
                points = []
                for t in np.linspace(0, 2 * np.pi, 500):
                    denom = 1 + np.sin(t)**2
                    if denom == 0:
                        continue  # avoid division by zero
                    x = (a * np.cos(t)) / denom
                    y = (a * np.cos(t) * np.sin(t)) / denom

                    # Convert to image coordinates
                    img_x = int(center_x + x)
                    img_y = int(center_y + y)
                    points.append((img_x, img_y))

                points.append(points[0])
                self.tracker.robot_list[-1].trajectory = points 
        
    

    def update_sensor_label(self):
        """
        Update the magnetic field sensor display labels.
        
        Called periodically by timer to refresh Bx, By, Bz sensor readings
        from the Arduino hall effect sensors.
        """
        # Replace this with your actual value source
        self.ui.bxlabel.setText("Bx: "+str(self.bx_sensor))
        self.ui.bylabel.setText("By: "+str(self.by_sensor))
        self.ui.bzlabel.setText("Bz: "+str(self.bz_sensor))

    

    def update_actions_frame(self, displayframe, cell_mask, robot_list, cell_list):
        """
        Main control loop that processes each video frame.
        
        This method:
        1. Reads hall effect sensor data from Arduino
        2. Executes path planning algorithms if enabled
        3. Runs control algorithms (orient/roll/push) if enabled
        4. Processes joystick input if enabled
        5. Updates robot and cell tracking data
        6. Saves data to Excel if recording
        7. Displays frame with overlays
        8. Sends commands to Arduino hardware
        
        Args:
            displayframe (np.ndarray): Current video frame to display
            cell_mask (np.ndarray): Binary mask of detected cells
            robot_list (list): List of tracked Robot objects
            cell_list (list): List of tracked Cell objects
        """
        #read hall effect sensor data from arduino
        sensor = self.arduino.receive()
        self.bx_sensor = -1 * round(sensor[0], 1)   #Bx sensor sign is switched
        self.by_sensor = round(sensor[1], 1)
        self.bz_sensor = round(sensor[2], 1)


        self.frame_number+=1


        #make path planner
        if self.path_planner_status == True:
            if len(robot_list) > 0:
                RRTtreesize = self.ui.RRTtreesizebox.value()
                self.tracker.robot_list[-1].trajectory = self.path_planner.run(robot_list, cell_mask, RRTtreesize)
                
                        


        if self.control_status == True:   
            arrivalthresh = self.ui.arrivalthreshbox.value()  
            
            if len(robot_list)>0:
                #orient algorithm option
                if self.ui.orientradio.isChecked():
                    displayframe, actions, stopped = self.control_robot.run_orient(displayframe, robot_list, arrivalthresh)
                    self.Bx, self.By, self.Bz, self.alpha, self.gamma, self.freq, self.psi, _  = actions   
                    #this is the auto acoustic opticmal frequency finder algorithm i designed
                    if self.ui.autoacousticbutton.isChecked():
                        self.acoustic_frequency = self.control_robot.find_optimal_acoustic_freq(robot_list, self.tracker.pixel2um)
                    # if stopped zero everything
                    if stopped == True:
                        self.apply_actions(False)
            

                #roll algorithm option
                elif self.ui.rollradio.isChecked():
                    displayframe, actions, stopped = self.control_robot.run_roll(displayframe, robot_list, arrivalthresh)
                    self.Bx, self.By, self.Bz, self.alpha, self.gamma, self.freq, self.psi, _  = actions    
                    
                    self.gamma = np.radians(self.ui.gammadial.value())
                    self.psi = np.radians(self.ui.psidial.value())
                    self.freq = self.ui.magneticfrequencydial.value()
                    if stopped == True:
                        self.apply_actions(False)
                    
                    
                #pushing algorithm option
                elif self.ui.pushradio.isChecked():
                    corridor_width = self.ui.corridorwidthbox.value()
                    approach_distance = self.ui.approachdistancebox.value()
                    spinning_freq = self.ui.spinningfreqbox.value()
                    pushingfreq = self.ui.magneticfrequencydial.value()

                    displayframe, actions = self.control_robot.run_push(displayframe, robot_list, cell_list, arrivalthresh, corridor_width,approach_distance, spinning_freq, pushingfreq, self.tracker.pixel2um)   
                    self.Bx, self.By, self.Bz, self.alpha, self.gamma, self.freq, self.psi, _  = actions  
                    self.psi = np.radians(self.ui.psidial.value())
            #zero option
            else:
                self.apply_actions(False)
                   


            
        #if joystick is on use the joystick though
        elif self.joystick_status == True:
            type, self.Bx, self.By, self.Bz, self.alpha, self.gamma, self.freq, self.psi, acoust = self.joystick_actions.run(self.joystick)
            self.psi = np.radians(self.ui.psidial.value())
            
            if acoust == 1:
                self.acoustic_frequency = self.ui.acousticfreq_spinBox.value()
            else:
                self.acoustic_frequency = 0

    

            if type == 1:
                self.gamma = np.radians(180)
                self.freq = self.ui.spinningfreqbox.value()
            
            elif type == 2:
                self.gamma = np.radians(0)
                self.freq = self.ui.spinningfreqbox.value()
            
            else:
                self.gamma = np.radians(self.ui.gammadial.value())
                if self.freq != 0:
                    self.freq = self.ui.magneticfrequencydial.value()
                    
                        
        
        
        elif self.manual_status == True:
            self.Bx = self.ui.manualfieldBx.value()/100
            self.By = self.ui.manualfieldBy.value()/100
            self.Bz = self.ui.manualfieldBz.value()/100
            
            """# X coil calibration
            if self.ui.manualfieldBx.value() == 0:  #0 mT
                self.Bx = 0  #0%  
            else:
                if self.ui.manualfieldBx.value() > 0:
                    self.Bx = 0.0511 * self.ui.manualfieldBx.value() **3    -    .2302   * self.ui.manualfieldBx.value() **2      +    .5577 *  self.ui.manualfieldBx.value()  + .0216
                elif self.ui.manualfieldBx.value() < 0:

                    self.Bx = -(0.0511 * (-self.ui.manualfieldBx.value()) **3    -    .2302   * (-self.ui.manualfieldBx.value()) **2      +    .5577 *  (-self.ui.manualfieldBx.value())  + .0216)


            # Y coil calibration
            if self.ui.manualfieldBy.value() == 0:  #0 mT
                self.By = 0  #0%  
            else:
                if self.ui.manualfieldBy.value() > 0:
                    self.By = .1863 *  self.ui.manualfieldBy.value()  + .07
                elif self.ui.manualfieldBy.value() < 0:
                    self.By = -(.1863 *  -self.ui.manualfieldBy.value()  + .07)


            # Z coil calibration
            if self.ui.manualfieldBz.value() == 0:  #0 mT
                self.Bz = 0  #0%  
            else:
                if self.ui.manualfieldBz.value() > 0:
                    self.Bz = 0.0003 * self.ui.manualfieldBz.value() **3    -    .0067   * self.ui.manualfieldBz.value() **2      +    .1027 *  self.ui.manualfieldBz.value()  + .023
                elif self.ui.manualfieldBz.value() < 0:
                    self.Bz = -(0.0003 * (-self.ui.manualfieldBz.value()) **3    -    .0067   * (-self.ui.manualfieldBz.value()) **2      +    .1027 *  (-self.ui.manualfieldBz.value())  + .023)"""

            


            self.freq = self.ui.magneticfrequencydial.value()
            self.gamma = np.radians(self.ui.gammadial.value())
            self.psi = np.radians(self.ui.psidial.value())
            self.alpha = np.radians(self.ui.alphadial.value())
           

            #ricochet conditions, too close to the x or y borders flip the conditions
            if self.ui.ricochet_effect_checkbox.isChecked():
                if len(self.tracker.robot_list) > 0:
                    for i in range(len(self.tracker.robot_list)):
                        bot = self.tracker.robot_list[i]
                        bot_pos_x = int(bot.position_list[-1][0])
                        bot_pos_y = int(bot.position_list[-1][1])
                        
                        vx = bot.velocity_list[-1][0]
                        vy = bot.velocity_list[-1][1] 
                        boundary = 200
                        #ricochet conditions, too close to the x or y borders
                        
                        if (bot_pos_x <= boundary and vx < 0) and (self.frame_number - self.ricochet_counter_x[-1] > 30):
                            vx = -vx
                            alpha = int(((np.degrees(np.arctan2(-vy,vx)) + 360) % 360))
                            self.ui.alphadial.setValue(alpha)
                            self.ricochet_counter_x.append(self.frame_number)

                        if  (bot_pos_x >= self.video_width - boundary and vx > 0) and (self.frame_number - self.ricochet_counter_x[-1] > 30):   
                            vx = -vx
                            alpha = int(((np.degrees(np.arctan2(-vy,vx)) + 360) % 360))
                            self.ui.alphadial.setValue(alpha)
                            self.ricochet_counter_x.append(self.frame_number)                  
                        
                        if (bot_pos_y <= boundary and vy < 0) and (self.frame_number - self.ricochet_counter_y[-1] > 30):
                            vy = -vy
                            alpha = int(((np.degrees(np.arctan2(-vy,vx)) + 360) % 360))
                            self.ui.alphadial.setValue(alpha)
                            self.ricochet_counter_y.append(self.frame_number)
            
                        if (bot_pos_y >= self.video_height - boundary and vy > 0) and (self.frame_number - self.ricochet_counter_y[-1] > 30):# and (self.frame_number - self.ricochet_counter_y[-1] > 30): #if the bot hits the top wall   
                            vy = -vy
                            alpha = int(((np.degrees(np.arctan2(-vy,vx)) + 360) % 360))
                            self.ui.alphadial.setValue(alpha)
                            self.ricochet_counter_y.append(self.frame_number)
                           
                    
                    
             
                        
            


        elif self.excel_actions_status == True and self.excel_actions_df is not None:            
            self.actions_counter +=1
            if self.actions_counter < self.excel_actions_df['Frame'].iloc[-1]:
                filtered_row = self.excel_actions_df[self.excel_actions_df['Frame'] == self.actions_counter]
                
                self.Bx = float(filtered_row["Bx"])
                self.By = float(filtered_row["By"])
                self.Bz = float(filtered_row["Bz"])
                self.alpha = float(filtered_row["Alpha"])
                self.gamma = float(filtered_row["Gamma"])
                self.freq = float(filtered_row["Rolling Frequency"])
                self.psi = float(filtered_row["Psi"])
                self.acoustic_freq = float(filtered_row["Acoustic Frequency"])
            
            else:
                self.excel_actions_status = False
                self.ui.apply_actions.setText("Apply")
                self.ui.apply_actions.setChecked(False)
                self.apply_actions(False)
            
        

        
        #DEFINE CURRENT ROBOT PARAMS TO A LIST
        if len(robot_list) > 0:
            self.robots = []
            for bot in robot_list:
                currentbot_params = [bot.frame_list[-1],
                                     bot.times[-1],
                                     bot.position_list[-1][0]* self.tracker.pixel2um,
                                     bot.position_list[-1][1]* self.tracker.pixel2um, 
                                     bot.velocity_list[-1][0]* self.tracker.pixel2um, 
                                     bot.velocity_list[-1][1]* self.tracker.pixel2um,
                                     bot.velocity_list[-1][2]* self.tracker.pixel2um,
                                     bot.acceleration_list[-1][0]* self.tracker.pixel2um,
                                     bot.acceleration_list[-1][1]* self.tracker.pixel2um,
                                     bot.acceleration_list[-1][2]* self.tracker.pixel2um,
                                     bot.blur_list[-1],
                                     bot.area_list[-1]* (self.tracker.pixel2um**2),
                                     self.tracker.pixel2um,
                                     [[x * self.tracker.pixel2um, y * self.tracker.pixel2um] for x, y in bot.trajectory]
                                    ]
            
                
                self.robots.append(currentbot_params)
           
      
        #DEFINE CURRENT CELL PARAMS TO A LIST
        if len(cell_list) > 0:
            self.cells = []
            for cell in cell_list:
                currentcell_params = [cell.frame_list[-1],
                                     cell.times[-1],
                                     cell.position_list[-1][0]* self.tracker.pixel2um,
                                     cell.position_list[-1][1]* self.tracker.pixel2um, 
                                     cell.velocity_list[-1][0]* self.tracker.pixel2um, 
                                     cell.velocity_list[-1][1]* self.tracker.pixel2um,
                                     cell.velocity_list[-1][2]* self.tracker.pixel2um,
                                     cell.blur_list[-1],
                                     cell.area_list[-1]* (self.tracker.pixel2um**2),
                                     self.tracker.pixel2um
                                    ]
                
                self.cells.append(currentcell_params)
        
        #DEFINE CURRENT MAGNETIC FIELD OUTPUT TO A LIST 
        self.actions = [self.frame_number, self.Bx, self.By, self.Bz, self.alpha, self.gamma, self.freq, self.psi, self.gradient_status,self.equal_field_status,
                        self.acoustic_frequency, self.bx_sensor, self.by_sensor, self.bz_sensor] 
       
        self.magnetic_field_list.append(self.actions)
        
        
        

        #IF SAVE STATUS THEN CONTINOUSLY SAVE THE CURRENT ROBOT PARAMS AND MAGNETIC FIELD PARAMS TO AN EXCEL ROWS
        if self.save_status == True:
            self.magnetic_field_sheet.append(self.actions)
            for (sheet, bot) in zip(self.robot_params_sheets,self.robots):
                sheet.append(bot[:-1])
            for (sheet, cell) in zip(self.cell_params_sheets,self.cells):
                sheet.append(cell[:-1])
        

        #also update robot info
        if len(self.robots) > 0:
            area = self.robots[-1][11]
            robot_diameter = round(np.sqrt(4*area/np.pi),1)
            self.ui.vellcdnum.display(int(self.robots[-1][6]))
            self.ui.blurlcdnum.display(int(self.robots[-1][10]))
            self.ui.accellcdnum.display(int(self.robots[-1][9]))
            self.ui.sizelcdnum.display(robot_diameter)

        
############################################################################################################################################################
        """Updates the image_label with a new opencv image"""
        if self.ui.toggledisplayvisualscheckbox.isChecked():
            if self.control_status == True or self.joystick_status == True or self.manual_status == True or self.excel_actions_status == True :
              
                rotatingfield = "alpha: {:.0f}, gamma: {:.0f}, psi: {:.0f}, freq: {:.0f}".format(np.degrees(self.alpha), np.degrees(self.gamma), np.degrees(self.psi), self.freq) #adding 90 to alpha for display purposes only
                
                cv2.putText(displayframe, rotatingfield,
                    (int(self.video_width / 1.8),int(self.video_height / 20)),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    fontScale=1.5, 
                    thickness=5,
                    color = (0, 0, 0),
                )
            
            acousticfreq = f'{self.acoustic_frequency:,} Hz'
            cv2.putText(displayframe, acousticfreq,
                (int(self.video_width / 8),int(self.video_height / 14)),
                cv2.FONT_HERSHEY_SIMPLEX,
                fontScale=1.5, 
                thickness=5,
                color = (0, 0, 0),
            )

        
        displayframe = self.handle_zoom(displayframe)
    

        rgb_image = cv2.cvtColor(displayframe, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
      
        bytes_per_line = ch * w
        convert_to_Qt_format = QtGui.QImage(rgb_image.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)
        p = convert_to_Qt_format.scaled(self.display_width, self.display_height, Qt.KeepAspectRatio)
        qt_img = QPixmap.fromImage(p)
       
        #update frame slider too
        self.ui.framelabel.setText("Frame:"+str(self.frame_number))
        if self.videopath !=0:
            self.ui.frameslider.setValue(self.tracker.framenum)
        
        self.ui.VideoFeedLabel.setPixmap(qt_img)


        #add frame to recording queue if record button is pressed
        if self.recorder and self.recorder.running:
            try:
                self.frame_queue.put_nowait(displayframe)
            except queue.Full:
                pass
      

        self.apply_actions(True)

           


        





    def apply_actions(self, status):
        """
        Apply current magnetic field and acoustic actions to hardware and simulator.
        
        Sends field parameters to Arduino for hardware control and updates
        the magnetic field simulator visualization.
        
        Args:
            status (bool): If True, apply current actions. If False, zero all outputs.
        """
        #the purpose of this function is to output the actions via arduino, 
        # show the actions via the simulator
        # and record the actions by appending the field_list
        
        if self.freq > 0:
            if self.ui.rollradio.isChecked() or self.ui.pushradio.isChecked():
                self.alpha = self.alpha + np.pi/2

        #zero output
        if status == False:
            self.manual_status = False
            self.Bx, self.By, self.Bz, self.alpha, self.gamma, self.freq, self.psi, self.acoustic_frequency = 0,0,0,0,0,0,0,0

        #output current actions to simulator

        self.simulator.Bx = self.Bx
        self.simulator.By = self.By
        self.simulator.Bz = self.Bz
        self.simulator.alpha = self.alpha
        self.simulator.gamma = self.gamma
        self.simulator.psi = self.psi
        self.simulator.freq = self.freq
        self.simulator.omega = 2 * np.pi * self.simulator.freq

        #send arduino commands
        self.arduino.send(self.Bx, self.By, self.Bz, self.alpha, self.gamma, self.freq, self.psi, self.gradient_status, self.equal_field_status, self.acoustic_frequency)


    def toggle_recording(self):
        """Toggle video recording on/off."""
        tracking_panel.toggle_recording(self)

    def start_data_record(self):
        """
        Initialize Excel workbook for data recording.
        
        Creates sheets for:
        - Magnetic field actions (Bx, By, Bz, angles, frequencies)
        - Robot tracking data (position, velocity, acceleration)
        - Cell tracking data
        """
        self.frame_number = 0
        self.tracker.framenum = 1
        self.tracker.start_time = time.time()
        self.tracker.time_stamp = 0
        
        self.output_workbook = openpyxl.Workbook()
            
        #create sheet for magneti field actions
        self.magnetic_field_sheet = self.output_workbook.create_sheet(title="Magnetic Field Actions")#self.output_workbook.active
        self.magnetic_field_sheet.append(["Frame","Bx (%)", "By (%)", "Bz (%)", "Alpha (rad)", "Gamma (rad)", "Rolling Frequency (Hz)", "Psi (rad)", "Gradient?","Equal Field?", "Acoustic Frequency (Hz)","Sensor Bx", "Sensor By", "Sensor Bz"])

        #create sheet for robot data
        self.robot_params_sheets = []
        for i in range(len(self.robots)):
            robot_sheet = self.output_workbook.create_sheet(title= "Robot {}".format(i+1))
            robot_sheet.append(["Frame","Time(s)","Pos X (um)", "Pos Y (um)", "Vel X (um/s)", "Vel Y (um/s)", "Vel Mag (um/s)", "Acc X (um/s2)", "Acc Y (um/s2)", "Acc Mag (um/s2)", "Blur", "Area (um^2)","pixel2um","Path X (um)", "Path Y (um)"])
            self.robot_params_sheets.append(robot_sheet)
        
        #create sheet for robot data
        self.cell_params_sheets = []
        for i in range(len(self.cells)):
            cell_sheet = self.output_workbook.create_sheet(title= "Cell {}".format(i+1))
            cell_sheet.append(["Frame","Time(s)","Pos X (um)", "Pos Y (um)", "Vel X (um/s)", "Vel Y (um/s)", "Vel Mag (um/s)", "Blur", "Area (um^2)","pixel2um"])
            self.cell_params_sheets.append(cell_sheet)

        #tell update_actions function to start appending data to the sheets
        self.save_status = True




    




    def stop_data_record(self):
        """
        Stop data recording and save Excel file.
        
        Finalizes the Excel workbook with trajectory data and saves to disk
        with timestamp filename.
        """
        #tell update_actions function to stop appending data to the sheets
        
        self.save_status = False
        file_path  = os.path.join(self.new_dir_path, self.output_file_name+".xlsx")
        print(self.output_file_name)
        
        #add trajectory to file after the fact
        if self.output_workbook is not None:
            if len((self.robot_params_sheets)) > 0:
                try:
                    for i in range(len((self.robot_params_sheets))):
                        for idx,(x,y) in enumerate(self.robots[i][-1]):
                            self.robot_params_sheets[i].cell(row=idx+2, column=14).value = x
                            self.robot_params_sheets[i].cell(row=idx+2, column=15).value = y
                except Exception:
                    pass
            
            #save and close workbook
            self.output_workbook.remove(self.output_workbook["Sheet"])
            self.output_workbook.save(file_path)

            self.output_workbook.close()
            self.output_workbook = None

    
    def savedata(self):
        """Toggle data recording on/off and handle UI updates."""
        if self.ui.savedatabutton.isChecked():
            self.ui.savedatabutton.setText("Stop")
            self.start_data_record()
            self.output_file_name = self.ui.videoNameLineEdit.text().strip()
            if not self.output_file_name:
                self.output_file_name = datetime.now().strftime("%Y-%m-%d_%H-%M-%S_%f")
        else:
            self.ui.savedatabutton.setText("Save Data")
            self.stop_data_record()
            



    def read_excel_actions(self):
        """
        Import pre-recorded magnetic field actions from Excel file.
        
        Opens file dialog for user to select Excel file containing
        frame-by-frame magnetic field parameters.
        """
        options = QFileDialog.Options()
        self.excel_file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls)", options=options)
        if self.excel_file_name:
            self.excel_actions_df = pd.read_excel(self.excel_file_name)
            
        
    def apply_excel_actions(self):
        """
        Toggle playback of imported Excel actions.
        
        When enabled, applies magnetic field parameters from Excel file
        synchronized to current frame number.
        """
        if self.ui.apply_actions.isChecked():
            self.excel_actions_status = True
            self.actions_counter = 0
            self.ui.apply_actions.setText("Stop")
        else:
            self.excel_actions_status = False
            self.ui.apply_actions.setText("Apply")
            self.apply_actions(False)
    
    


    
    def toggle_simulation(self):
        """Toggle magnetic field simulator visualization on/off."""
        if self.ui.simulationbutton.isChecked():
            self.simulator.start()
            self.tbprint("Simulation Off")
            self.ui.simulationbutton.setText("Simulation Off")
        else:
            self.simulator.stop()
            self.tbprint("Simulation On")
            self.ui.simulationbutton.setText("Simulation On")
   
    
    
    def toggle_control_status(self): 
        """
        Toggle autonomous control algorithms on/off.
        
        When enabled, control algorithms (orient/roll/push) run automatically
        based on tracked robot positions and user-defined waypoints.
        """
        if self.ui.controlbutton.isChecked():
            self.control_robot.reset()
            self.control_status = True
            self.ui.controlbutton.setText("Stop")
            self.tbprint("Control On: {} Hz".format(self.acoustic_frequency))
        else:
            self.control_status = False
            self.ui.controlbutton.setText("Control")
            self.tbprint("Control Off")
            self.apply_actions(False)
          
         
            
            
    

    def toggle_joystick_status(self):
        """
        Toggle joystick control mode on/off.
        
        Allows manual control of magnetic fields using a connected USB joystick.
        Checks for joystick availability before enabling.
        """
        if pygame.joystick.get_count() != 0:
            if self.ui.joystickbutton.isChecked():
                self.joystick_status = True
                self.ui.joystickbutton.setText("Stop")
                self.tbprint("Joystick On")
            else:
                self.joystick_status = False
                self.ui.joystickbutton.setText("Joystick")
                self.tbprint("Joystick Off")
                self.apply_actions(False)
        else:
            self.tbprint("No Joystick Connected...")

    def toggle_autoacoustic(self):
        """
        Toggle automatic acoustic frequency optimization.
        
        When enabled, automatically calculates optimal acoustic frequency
        based on robot size and material properties.
        """
        if self.tracker is not None:
            if self.ui.autoacousticbutton.isChecked():
                self.autoacousticstatus = True
                self.ui.led.setStyleSheet("\n"
"                background-color: rgb(0, 255, 0);\n"
"                border-style: outset;\n"
"                border-width: 3px;\n"
"                border-radius: 12px;\n"
"                border-color: rgb(0, 255, 0);\n"
"         \n"
"                padding: 6px;")
            else:
                self.autoacousticstatus = False
                self.acoustic_frequency = 0
                self.ui.led.setStyleSheet("\n"
"                background-color: rgb(255, 0, 0);\n"
"                border-style: outset;\n"
"                border-width: 3px;\n"
"                border-radius: 12px;\n"
"                border-color: rgb(255, 0, 0);\n"
"         \n"
"                padding: 6px;")
                

                

    def get_acoustic_frequency(self):
        """Read acoustic frequency value from UI spin box."""
        if self.ui.applyacousticbutton.isChecked():
            self.acoustic_frequency = self.ui.acousticfreq_spinBox.value()
            #self.tbprint("Control On: {} Hz".format(self.acoustic_frequency))
            self.apply_acoustic()
        
    
    def apply_acoustic(self):
        """
        Apply or stop acoustic stimulation.
        
        Updates LED indicator and sends frequency to Arduino.
        """
        if self.ui.applyacousticbutton.isChecked():
            self.ui.applyacousticbutton.setText("Stop")
            #self.tbprint("Control On: {} Hz".format(self.acoustic_frequency))
            self.acoustic_frequency = self.ui.acousticfreq_spinBox.value()
            #self.apply_actions(True)
            self.ui.led.setStyleSheet("\n"
"                background-color: rgb(0, 255, 0);\n"
"                border-style: outset;\n"
"                border-width: 3px;\n"
"                border-radius: 12px;\n"
"                border-color: rgb(0, 255, 0);\n"
"         \n"
"                padding: 6px;")
        
        else:
            self.ui.applyacousticbutton.setText("Apply")
            #self.tbprint("Acoustic Module Off")
            self.acoustic_frequency = 0
            self.ui.led.setStyleSheet("\n"
"                background-color: rgb(255, 0, 0);\n"
"                border-style: outset;\n"
"                border-width: 3px;\n"
"                border-radius: 12px;\n"
"                border-color: rgb(255, 0, 0);\n"
"         \n"
"                padding: 6px;")
            #self.apply_actions(False)
       
        

    def tbprint(self, text):
        """
        Print text to the UI console/textbox.
        
        Args:
            text (str): Message to display in console
        """
        #print to textbox
        self.ui.plainTextEdit.appendPlainText("$ "+ text)
    

    def convert_coords(self,pos):
        """
        Convert mouse position from display coordinates to video coordinates.
        
        Accounts for scaling between video resolution and display widget size.
        
        Args:
            pos (QPoint): Mouse position in display coordinates
            
        Returns:
            tuple: (x, y) position in video coordinates
        """
        #need a way to convert the video position of mouse to the actually coordinate in the window
        newx = int(pos.x() * (self.video_width / self.display_width)) 
        newy = int(pos.y() * (self.video_height / self.display_height))
        return newx, newy
    
    
    def keyPressEvent(self, event):
        # Check if the event type is QEvent.KeyPress
        if event.type() == QEvent.KeyPress:
            # Check if the key pressed is 'c'
            if event.key() == Qt.Key_C:
                print('c')
        # Call the base class method for default processing
        return super().keyPressEvent(event)
 

    

    def eventFilter(self, object, event):
        """
        Filter and handle mouse events on the video display.
        
        Processes clicks, drags, and wheel events for robot selection,
        waypoint creation, and zoom control.
        
        Args:
            object (QObject): Object that received the event
            event (QEvent): Event to process
            
        Returns:
            bool: True if event was handled, False otherwise
        """
        return tracking_panel.eventFilter(self, object, event)
            
            
        
        

    

    def update_croppedimage(self, frame, recoreded_frame):
        """
        Update the cropped robot view display.
        
        Shows zoomed-in view of selected robot and optionally records it.
        
        Args:
            frame (np.ndarray): Cropped frame to display
            recoreded_frame (np.ndarray): Frame to record if recording is active
        """
        """Updates the cropped image_label with a new cropped opencv image"""
        
        rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
        
        bytes_per_line = ch * w
        convert_to_Qt_format = QtGui.QImage(rgb_image.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)
        p = convert_to_Qt_format.scaled(310, 310, Qt.KeepAspectRatio)
        qt_cimg = QPixmap.fromImage(p)
        self.ui.CroppedVideoFeedLabel.setPixmap(qt_cimg)
        
        #recored the robots suroundings
        if self.croppedresult is not None:
            self.croppedresult.write(recoreded_frame)

    

    def croppedrecordfunction(self):
        """Toggle recording of cropped robot view."""
        if self.cap is not None:
            if self.ui.croppedrecordbutton.isChecked():
                self.ui.croppedrecordbutton.setText("Stop")
                self.tbprint("Start Record")
                self.date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S_%f")
                file_path  = os.path.join(self.new_dir_path, self.date+".mp4")
                self.croppedresult = cv2.VideoWriter(
                    file_path,
                    cv2.VideoWriter_fourcc(*"mp4v"),
                    int(self.videofps),    
                    (200, 200), ) 
                #start recording magnetic field and tracking data
                self.start_data_record()
            
            else:
                self.ui.croppedrecordbutton.setText("Record")
                if self.croppedresult is not None:
                    self.croppedresult.release()
                    self.croppedresult = None
                    self.tbprint("End Record, Data Saved")
                #stop and save the data when the record is over.
                self.stop_data_record()
    
         
    

    
    def setFile(self):
        """
        Initialize video capture source.
        
        Attempts to connect to FLIR camera using EasyPySpin, falls back to
        default camera if unavailable. Configures display dimensions and frame rate.
        """
        if self.videopath == 0:
            try:
                self.cap  = EasyPySpin.VideoCapture(0)
   
                self.cap.set(cv2.CAP_PROP_AUTO_WB, True)
                self.cap.set(cv2.CAP_PROP_FPS, 24)
                self.tbprint("Connected to FLIR Camera")

                if not self.cap.isOpened():
                    self.cap  = cv2.VideoCapture(0) 
                    self.tbprint("No EasyPySpin Camera Available")
            
            except Exception:
                self.cap  = cv2.VideoCapture(0) 
                self.tbprint("No EasyPySpin Camera Available")
                
                
            self.ui.pausebutton.hide()
            self.ui.leftbutton.hide()
            self.ui.rightbutton.hide()
            self.ui.frameslider.hide()
        else:
            self.cap  = cv2.VideoCapture(self.videopath)
            self.ui.pausebutton.show()
            self.ui.leftbutton.show()
            self.ui.rightbutton.show()
        
        self.video_width = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        self.video_height = int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        self.videofps = int(self.cap.get(cv2.CAP_PROP_FPS))
        self.tbprint("Width: {}  --  Height: {}  --  Fps: {}".format(self.video_width,self.video_height,self.videofps))

        self.aspectratio = (self.video_width / self.video_height)

        self.resize_widgets()        

        if self.videopath == 0:
            self.ui.robotsizeunitslabel.setText("um")
            self.ui.robotvelocityunitslabel.setText("um/s")
        else:
            self.ui.robotsizeunitslabel.setText("px")
            self.ui.robotvelocityunitslabel.setText("px/s")
            self.totalnumframes = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
            self.tbprint("Total Frames: {} ".format(self.totalnumframes))
            self.ui.frameslider.setGeometry(QtCore.QRect(10, self.display_height+12, self.display_width, 20))
            self.ui.frameslider.setMaximum(self.totalnumframes)
            self.ui.frameslider.show()
        
        #if self.ui.recordbutton.isChecked():
            #self.recordfunction()

        #if not self.ui.startbutton.isChecked(): #clear the pixmap
        self.ui.VideoFeedLabel.setPixmap(QtGui.QPixmap())
        


    def selectFile(self):
        """
        Open file dialog to select video file for playback.
        
        Allows loading recorded videos instead of live camera feed.
        """
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.ReadOnly
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Open File", "", "All Files (*);;Text Files (*.txt);;Python Files (*.py)", options=options)

        if file_path:
            self.videopath = file_path
            file_info = QtCore.QFileInfo(file_path)
            file_name = file_info.fileName()
            self.ui.choosevideobutton.setText(file_name)
            self.tbprint(file_name)
        else:
            self.videopath = 0
            self.ui.choosevideobutton.setText("Live")
            self.tbprint("Using Video Camera")
        
        self.setFile()
        
    
    
    def populate_serial_ports(self):
        """Scan and populate dropdown with available serial ports for Arduino."""
        ports = list_ports.comports()
        if len(ports) > 0:
            self.ui.arduino_portbox.clear()
            for port in ports:
                self.ui.arduino_portbox.addItem(port.device)
            self.arduino_port = port.device
        
        else:
            self.arduino_port = None

        

    def handle_port_change(self, selected_port):
        """
        Handle Arduino serial port selection change.
        
        Args:
            selected_port (str): Name of selected serial port
        """
        self.arduino_port = selected_port
        



    def start(self):
        """Start video capture and tracking loop."""
        tracking_panel.start(self)

    def showmask(self):
        """Toggle display between original video and segmentation mask."""
        if self.tracker is not None:
            if self.ui.maskbutton.isChecked():
                self.ui.maskbutton.setText("Original")
                self.tracker.mask_flag = True
            else:
                self.ui.maskbutton.setText("Mask")
                self.tracker.mask_flag = False
    
    def showcroppedoriginal(self):
        """Toggle cropped view between original and mask display."""
        if self.tracker is not None:
            if self.ui.croppedmasktoggle.isChecked():
                self.ui.croppedmasktoggle.setText("Mask")
                self.tracker.croppedmask_flag = False
            else:
                self.ui.croppedmasktoggle.setText("Original")
                self.tracker.croppedmask_flag = True


    def spinbox_alphachanged(self):
        """Sync alpha angle dial with spinbox value."""
        self.ui.alphadial.setValue(self.ui.alphaspinBox.value())
    
    def dial_alphachanged(self):
        """Sync alpha angle spinbox with dial value."""
        self.ui.alphaspinBox.setValue(self.ui.alphadial.value())

    def gradientcommand(self):
        """Toggle magnetic field gradient mode on/off."""
        self.gradient_status = int(self.ui.gradient_status_checkbox.isChecked())

    def equalfieldcommand(self):
        """Toggle equal magnetic field mode on/off."""
        self.equal_field_status = int(self.ui.equal_field_checkbox.isChecked())

    def get_objective(self):
        """Read microscope objective magnification from UI."""
        if self.tracker is not None:
            self.tracker.objective = self.ui.objectivebox.value()

    def get_exposure(self):
        """Read camera exposure time from UI and apply to camera."""
        if self.tracker is not None:
            self.tracker.exposure = self.ui.exposurebox.value()
    


    def invertmaskcommand(self):
        """Toggle mask inversion for tracking bright or dark objects."""
        if self.tracker is not None:
            self.ui.maskinvert_checkBox.setText("Invert Mask: " + str(self.ui.maskinvert_checkBox.isChecked()))
            self.tracker.maskinvert = self.ui.maskinvert_checkBox.isChecked()

    def pause(self):
        """Pause or resume video playback (video files only)."""
        if self.videopath != 0:
            if self.ui.pausebutton.isChecked():
                self.tracker._play_flag = False
                self.ui.pausebutton.setText("Play")
              
            else:#play
                self.tracker._play_flag = True
                self.ui.pausebutton.setText("Pause")
                
    def frameright(self):
        """Advance video by one frame (video files only)."""
        if self.videopath != 0:
            self.tracker.framenum+=1
            self.ui.frameslider.setValue(self.tracker.framenum)
            self.ui.framelabel.setText("Frame:"+str(self.tracker.framenum))

    def frameleft(self):
        """Rewind video by one frame (video files only)."""
        if self.videopath != 0:
            self.tracker.framenum-=1
            self.ui.frameslider.setValue(self.tracker.framenum)
            self.ui.framelabel.setText("Frame:"+str(self.tracker.framenum))

    
    
    
    def get_manual_bfieldbuttons(self):
        """Toggle manual magnetic field control mode."""
        if self.ui.manualapplybutton.isChecked():
            self.manual_status = True
            self.ui.manualapplybutton.setText("Stop")
        else:
            self.ui.manualapplybutton.setText("Apply")
            self.apply_actions(False)
    


       
    def get_slider_vals(self):
        """
        Read all UI slider and spinbox values and apply to tracker.
        
        Updates tracking parameters including mask thresholds, dilation,
        blur, crop sizes, and control parameters.
        """
        memory = self.ui.memorybox.value()
        magneticfreq = self.ui.magneticfrequencydial.value()
        gamma = self.ui.gammadial.value()
        psi = self.ui.psidial.value()
        #alpha = self.ui.alphaspinBox.value()
        
        robotlower = self.ui.robotmasklowerbox.value() 
        robotupper = self.ui.robotmaskupperbox.value()
        robotdilation = self.ui.robotmaskdilationbox.value() 
        robotmaskblur = self.ui.robotmaskblurbox.value()
        robotcrop_length = self.ui.robotcroplengthbox.value()
        
        celllower = self.ui.cellmasklowerbox.value() 
        cellupper = self.ui.cellmaskupperbox.value()
        celldilation = self.ui.cellmaskdilationbox.value() 
        cellmaskblur = self.ui.cellmaskblurbox.value()
        cellcrop_length = self.ui.cellcroplengthbox.value()
        

        if self.tracker is not None: 
            self.tracker.memory = memory
          

          
            self.tracker.robot_mask_lower = robotlower
            self.tracker.robot_mask_upper = robotupper
            self.tracker.robot_mask_dilation = robotdilation
            self.tracker.robot_mask_blur = robotmaskblur
            self.tracker.robot_crop_length = robotcrop_length
            
    
            self.tracker.cell_mask_lower = celllower
            self.tracker.cell_mask_upper = cellupper
            self.tracker.cell_mask_dilation = celldilation
            self.tracker.cell_mask_blur = cellmaskblur
            self.tracker.cell_crop_length = cellcrop_length
            

        self.ui.gammalabel.setText("Gamma: {}".format(gamma))
        self.ui.psilabel.setText("Psi: {}".format(psi))
        #self.ui.rollingfrequencylabel.setText("Freq: {}".format(magneticfreq))

         
        
    def resetparams(self):
        """Reset all tracking and control parameters to default values."""
        self.ui.robotmasklowerbox.setValue(0)
        self.ui.robotmaskupperbox.setValue(128)
        self.ui.robotmaskdilationbox.setValue(0)
        self.ui.robotmaskblurbox.setValue(0)
        self.ui.robotcroplengthbox.setValue(40)

        self.ui.cellmasklowerbox.setValue(0)
        self.ui.cellmaskupperbox.setValue(128)
        self.ui.cellmaskdilationbox.setValue(0)
        self.ui.cellmaskblurbox.setValue(0)
        self.ui.cellcroplengthbox.setValue(40)
        

    
        self.ui.memorybox.setValue(15)
        self.ui.RRTtreesizebox.setValue(25)
        self.ui.arrivalthreshbox.setValue(100)
        self.ui.gammadial.setSliderPosition(90)
        self.ui.psidial.setSliderPosition(90)
        self.ui.magneticfrequencydial.setValue(10)
        self.ui.acousticfreq_spinBox.setValue(1000000)
        self.ui.objectivebox.setValue(10)
        self.ui.exposurebox.setValue(5000)
        

    def resizeEvent(self, event):
        """
        Handle window resize events.
        
        Args:
            event (QResizeEvent): Resize event with new dimensions
        """
        windowsize = event.size()
        self.window_width = windowsize.width()
        self.window_height = windowsize.height()
        self.resize_widgets()
 
    def resize_widgets(self):
        """
        Adjust widget sizes and positions based on current window size.
        
        Maintains aspect ratio of video display and scales UI elements proportionally.
        """
        self.display_height = int(self.window_height*self.displayheightratio) #keep this fixed, changed the width dpending on the aspect ratio
        self.framesliderheight = int(self.window_height*self.framesliderheightratio)
        self.textheight = int(self.window_height*self.textheightratio)
        self.tabheight = self.window_height*self.tabheightratio
        self.display_height = int(self.window_height*self.displayheightratio) #keep this fixed, changed the width dpending on the aspect ratio
        self.framesliderheight = int(self.window_height*self.framesliderheightratio)
        self.textheight = int(self.window_height*self.textheightratio)
        self.tabheight = self.window_height*self.tabheightratio

        self.display_width = int(self.display_height * self.aspectratio)

        self.ui.VideoFeedLabel.setGeometry(QtCore.QRect(10,  5,                       self.display_width,     self.display_height))
        self.ui.frameslider.setGeometry(QtCore.QRect(10,    self.display_height+12,   self.display_width,     self.framesliderheight))
        self.ui.plainTextEdit.setGeometry(QtCore.QRect(10,  self.display_height+20+self.framesliderheight,   self.display_width,     self.textheight))

        #self.ui.tabWidget.setGeometry(QtCore.QRect(12,  6,  260 ,     self.tabheight))

    def handle_zoom(self, frame):
        """
        Apply digital zoom to a region of the video frame.
        
        Crops and scales a region around the specified zoom coordinates,
        then overlays it back onto the original frame.
        
        Args:
            frame (np.ndarray): Video frame to zoom
            
        Returns:
            np.ndarray: Frame with zoomed region
        """
        if self.zoomscale > 1:
            x = self.zoom_x
            y = self.zoom_y
            w = 300
            h = 300
            angle = 0
            
            # step 1: cropped a frame around the coord you wont to zoom into
            if y-w < 0 and x-h < 0:
                zoomedframe = frame[0:y+h , 0:x+w]
                cv2.rectangle(frame, (0, 0), (x + w, y + h), (0, 255, 0), 2)
                warpx = x
                warpy = y
            elif x-w < 0:
                zoomedframe = frame[y-h:y+h , 0:x+w] 
                cv2.rectangle(frame, (0, y-h), (x + w, y + h), (0, 255, 0), 2)
                warpx = x
                warpy = h
            elif y-h < 0:
                zoomedframe = frame[0:y+h , x-w:x+w]
                cv2.rectangle(frame, (x-w, 0), (x + w, y + h), (0, 255, 0), 2)
                warpx = w
                warpy = y
            else:
                zoomedframe = frame[y-h:y+h , x-w:x+w] 
                cv2.rectangle(frame, (x-w, y-h), (x + w, y + h), (0, 255, 0), 2)
                warpx = w
                warpy = h   
            
            # step 2: zoom into the zoomed frame a certain zoom amount
            rot_mat = cv2.getRotationMatrix2D((warpx,warpy), angle, self.zoomscale)
            zoomedframe = cv2.warpAffine(zoomedframe, rot_mat, zoomedframe.shape[1::-1], flags=cv2.INTER_LINEAR)

            #step 3: replace the original cropped frame with the new zoomed in cropped frame
            if y-h < 0 and x-w < 0:
                frame[0:y+h , 0:x+w] =  zoomedframe
            elif x-w < 0:
                frame[y-h:y+h , 0:x+w] =  zoomedframe
            elif y-h < 0:
                frame[0:y+h , x-w:x+w] =  zoomedframe
            else:
                frame[y-h:y+h , x-w:x+w] =  zoomedframe


        
        return frame

    def closeEvent(self, event):
        """
        Handle application close event.
        
        Called when X button is pressed

        Ensures proper cleanup of threads, hardware connections, and data files.
        
        Args:
            event (QCloseEvent): Close event
        """
        
        if self.tracker is not None:
            self.tracker.stop()
        #self.recorder.stop()
        
        self.simulator.stop()
        
        if self.arduino is not None:
            self.arduino.close()
            self.apply_actions(False)
